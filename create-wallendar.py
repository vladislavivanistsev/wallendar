from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
import calendar

def recolor_calendar(ws):
    # Define colors
    workday_color  = "E0FAE9"
    saturday_color = "C2F6D3"
    sunday_color   = "A5F2BF"
    border_color   = "22DD60"

    # Recolor every odd month
    for month_idx in range(2, 14, 2):  # Start from 2 to skip the header row, increment by 2 for odd months
        for col in range(2, ws.max_column + 1):  # Skip the month abbreviation column
            cell = ws.cell(row=month_idx, column=col)
            if cell.value is not None:  # Check if the cell is not empty
                cell.fill = PatternFill(start_color=workday_color, fill_type="solid")
            else:  # Keep the empty cell white or transparent
                cell.fill = PatternFill(start_color="FFFFFF", fill_type="solid")

    # Find and recolor Saturday (Sat) and Sunday (Sun) columns
    for col in range(2, ws.max_column + 1):  # Start from 2 to skip the month abbreviation column
        if ws.cell(row=1, column=col).value == "Sa":  # Check if the header is Saturday
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:  # Check if the cell is not empty
                    cell.fill = PatternFill(start_color=saturday_color, fill_type="solid")
                else:  # Keep the empty cell white or transparent
                    cell.fill = PatternFill(start_color="FFFFFF", fill_type="solid")
        elif ws.cell(row=1, column=col).value == "Su":  # Check if the header is Sunday
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:  # Check if the cell is not empty
                    cell.fill = PatternFill(start_color=sunday_color, fill_type="solid")
                else:  # Keep the empty cell white or transparent
                    cell.fill = PatternFill(start_color="FFFFFF", fill_type="solid")

    # Recolor first and last column, first and last row
    last_col = ws.max_column
    last_row = ws.max_row
    for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=last_col):
        for cell in row:
            if cell.row in [1, last_row] or cell.column in [1, last_col]:
                cell.fill = PatternFill(start_color=border_color, fill_type="solid")


def create_excel_wallendar(year):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Calendar {year}"

    # Define colors
    border_color = "22DD60"

    # Define the border style
    border = Border(left=Side(style='thin', color=border_color),
                    right=Side(style='thin', color=border_color),
                    top=Side(style='thin', color=border_color),
                    bottom=Side(style='thin', color=border_color))

    # Abbreviations for the months
    month_abbr = [calendar.month_abbr[m] for m in range(1, 13)]

    # Days of the week headers
    days_of_week = ['Mo', 'Tu', 'We', 'Th', 'Fr', 'Sa', 'Su']

    last_col = 0  # Track the last filled column

    # Generate the calendar for each month
    for month_idx, month in enumerate(month_abbr, start=1):
        # Month abbreviation in column A
        cell = ws.cell(row=month_idx + 1, column=1)
        cell.value = month
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(name="Ubuntu", color="000000")
        cell.border = border

        # Calculate the first weekday of the month and the number of days
        first_weekday, num_days = calendar.monthrange(year, month_idx)
        start_col = 2 + first_weekday  # +2 to account for the month column and 0-indexing

        # Fill in the days of the month
        for day in range(1, num_days + 1):
            col = start_col + day - 1
            cell = ws.cell(row=month_idx + 1, column=col)
            cell.value = day
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(name="Ubuntu", color="000000")
            cell.border = border
            last_col = max(last_col, col)  # Update the last filled column

    # Apply special formatting to the first and last columns, and first and last rows
    for i in range(1, last_col + 2):
        for row in [1, 14]:
            cell = ws.cell(row=row, column=i)
            cell.font = Font(name="Ubuntu", color="FFFFFF")
            cell.border = border
        for row in range(2, 14):
            for col in [1, last_col + 1]:
                cell = ws.cell(row=row, column=col)
                cell.font = Font(name="Ubuntu", color="FFFFFF")
                cell.border = border

    # Fill row 1 and row 14 with weekday abbreviations up to the last column
    for i in range(2, last_col + 1):  # Starting from column 2
        day_of_week = days_of_week[(i - 2) % 7]
        for row in [1, 14]:
            cell = ws.cell(row=row, column=i)
            cell.value = day_of_week
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(name="Ubuntu", color="FFFFFF")
            cell.border = border

    # In the last column + 1, from rows 2 to 13, repeat the month abbreviations
    for month_idx, month in enumerate(month_abbr, start=1):
        cell = ws.cell(row=month_idx + 1, column=last_col + 1)
        cell.value = month
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(name="Ubuntu", color="FFFFFF")
        cell.border = border

    # Remove all unfilled columns and rows from the Excel table
    ws.delete_cols(last_col + 3, ws.max_column)
    ws.delete_rows(15, ws.max_row)

    # After setting up the calendar and before saving the workbook, call recolor_calendar
    recolor_calendar(ws)

    # Save the workbook
    wb.save(f"wall-calendar-{year}.xlsx")

create_excel_wallendar(2025)

# Adjust the cells size in excel.
# Remove margins and set Landscape A2
# Change the font size and orientation
# Fit to the printing area, export as pdf
